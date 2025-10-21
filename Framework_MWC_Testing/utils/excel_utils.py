import os
import json
import csv
import pandas as pd
from typing import List, Dict, Any, DefaultDict
from collections import defaultdict
from openpyxl import load_workbook

# ========== HÀM TIỆN ÍCH CƠ BẢN ==========
def ensure_dir(path: str):
    """Tạo thư mục nếu chưa tồn tại."""
    os.makedirs(path, exist_ok=True)

def _norm(value) -> str:
    """Chuẩn hóa dữ liệu đầu vào (loại bỏ None, NaN, null)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ["nan", "none", "null"]:
        return ""
    return s

def _norm_key(s) -> str:
    """Đưa header về chữ thường."""
    return _norm(s).lower()


# ========== ĐỌC EXCEL ==========
def load_sheet(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_norm_key(h) for h in rows[0]]
    data = []
    for r in rows[1:]:
        if not r:
            continue
        row = {header[i]: _norm(r[i]) for i in range(min(len(header), len(r)))}
        if row.get("testcase"):
            data.append(row)

    print(f"[INFO] Loaded {len(data)} dòng từ sheet '{sheet_name}' ({os.path.basename(path)})")
    return data


# ========== ĐỌC CSV ==========
def load_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        data = []
        for row in reader:
            clean = {k.strip().lower(): _norm(v) for k, v in row.items()}
            if clean.get("testcase"):
                data.append(clean)
    print(f"[INFO] Loaded {len(data)} dòng từ CSV: {os.path.basename(path)}")
    return data


# ========== ĐỌC JSON ==========
def load_json(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    data = []
    for item in raw:
        clean = {k.strip().lower(): _norm(v) for k, v in item.items()}
        if clean.get("testcase"):
            data.append(clean)
    print(f"[INFO] Loaded {len(data)} dòng từ JSON: {os.path.basename(path)}")
    return data


# ========== HÀM CHUNG ==========
def load_data(path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return load_sheet(path, sheet_name or "Sheet1")
    elif ext == ".csv":
        return load_csv(path)
    elif ext == ".json":
        return load_json(path)
    else:
        raise ValueError(f"Không hỗ trợ định dạng: {ext}")


# ========== GHI KẾT QUẢ ==========
class ResultBook:
    def __init__(self, out_dir: str, file_name: str = "ResultsData.xlsx"):
        ensure_dir(out_dir)
        self.path = os.path.join(out_dir, file_name)
        self._sheets: DefaultDict[str, list[dict]] = defaultdict(list)

    def add_row(self, sheet: str, row: Dict[str, Any]):
        self._sheets[sheet].append(row)

    def save(self):
        existing = {}
        if os.path.exists(self.path):
            try:
                xls = pd.ExcelFile(self.path, engine="openpyxl")
                for s in xls.sheet_names:
                    existing[s] = xls.parse(s)
            except Exception:
                pass

        with pd.ExcelWriter(self.path, engine="openpyxl", mode="w") as writer:
            for s, df_old in existing.items():
                if s not in self._sheets:
                    df_old.to_excel(writer, sheet_name=s, index=False)
            for s, rows in self._sheets.items():
                pd.DataFrame(rows).to_excel(writer, sheet_name=s, index=False)

        print(f"[RESULT SAVED]: {self.path}")
        return self.path
